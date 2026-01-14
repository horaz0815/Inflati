#!/usr/bin/env python3
"""
Script to create an Excel file with names from the attendance lists
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# List of names from both images
names = [
    # From first image
    ("GMOSER", "Michael"),
    ("PRIBASNIG", "Thomas"),
    ("KLEIN", "Rudolf"),
    ("GEIGER", "Christian"),
    ("GIRITZER", "Markus"),
    ("JURIS", "Markus"),
    ("HORACEK", "Michael"),
    ("NOESTLER", "Stefan"),
    ("FUCHS", "Martin"),
    ("ALLRAM", "Lukas"),
    ("ASCHAUER", "Thomas"),
    ("BERLAKOVITS", "Stefan"),
    ("HÖRMANN", "Armin"),
    ("RIEPL", "Alexander"),
    ("MITTERBÖCK", "Lukas"),
    ("ZEHETMAYER", "Fabian"),
    ("KUGLER", "Daniel"),
    ("URBAN", "Patrick"),
    ("NEKOLA", "Maximilian"),
    ("KONSTANDINOV", "Nikola"),
    ("HOHENAUER", "Matthias"),
    # From second image
    ("KAPPL", "Christian"),
    ("WYHLIDAL", "Jürgen"),
    ("HAAS", "Christopher"),
    ("GUTSCHER", "Christiane"),
    ("BEYRER", "Jürgen"),
    ("FORSTNER", "Patrick"),
    ("WEIß", "Michael"),
    ("PUSCHNIGG", "Andreas"),
    ("TAUCHER", "Armin"),
    ("GRUBER", "Richard"),
    ("JANECZEK", "Raphael"),
    ("FLEISCHHACKER", "Stefan"),
    ("KLEMENS", "Florian"),
    ("HOCHÖRTLER", "Emanuel"),
    ("PETER", "Simon"),
    ("STIEG", "Philip"),
    ("KOLLMANN", "Alexander"),
    ("MAIR", "Wolfgang"),
    ("EISNER", "Alfred"),
    ("BAUER", "Christoph"),
    ("HOFMAIER", "Patrick"),
    ("PFEFFER", "Michael"),
    ("CZERMAK", "Andre"),
]

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Namensliste"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
alignment_center = Alignment(horizontal="center", vertical="center")

# Add headers
ws['A1'] = "Nr."
ws['B1'] = "Nachname"
ws['C1'] = "Vorname"

# Apply header styling
for col in ['A1', 'B1', 'C1']:
    ws[col].fill = header_fill
    ws[col].font = header_font
    ws[col].alignment = alignment_center

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20

# Add data
for idx, (lastname, firstname) in enumerate(names, start=1):
    row = idx + 1
    ws[f'A{row}'] = idx
    ws[f'B{row}'] = lastname
    ws[f'C{row}'] = firstname

    # Center align the number column
    ws[f'A{row}'].alignment = alignment_center

# Save the workbook
wb.save("Namensliste.xlsx")
print(f"Excel-Datei erstellt: Namensliste.xlsx")
print(f"Anzahl der Namen: {len(names)}")
